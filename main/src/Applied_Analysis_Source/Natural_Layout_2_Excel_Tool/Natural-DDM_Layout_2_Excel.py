import re
import sys
import os
import glob


import tqdm
import pandas
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side


EXCEL_TITLE = ["No", "DBIO", "FILE-ID", "View名", "レベル", "グループ", "物理名", "論理名", "フォーマット", "長さ", "圧縮", "ディスクリプタ"]
TYL, DB, NAME, F, LENG, S, D, REMARKS = 0, 1, 2, 3, 4, 5, 6, 7

PATTERN_COMMENT = re.compile(r"^\s*\*")
PATTERN_F = re.compile(r"^1NEXT\s*L\s*F\s*(\S+?)[\s]")
PATTERN_DB_FILE = re.compile(r"^1\s*DB:\s*(\S+?)\s*FILE:\s*(\S+?)[\s]")
PATTERN_LEN_LINE = re.compile(r"^\s*-")


def write_excel(output_path: str, result_list: list, column: list, format=True, index_flg=False) -> None:
    pandas.DataFrame(result_list, columns=column).to_excel(output_path, index=index_flg, engine="xlsxwriter")
    if format:
        formatter_excel(output_path)


def formatter_excel(excel_path) -> None:
    print(f"Formatting Excel: {os.path.basename(excel_path)}...")
    work_book = openpyxl.load_workbook(excel_path)
    for work_sheet_name in work_book.sheetnames:
        print(f" Sheet -> {work_sheet_name}...")
        work_sheet = work_book[work_sheet_name]
        max_row = work_sheet.max_row
        max_column = work_sheet.max_column
        max_column_len = [0] * (max_column + 1)
        # 设置边框
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )
        # 设置字体
        content_font = Font(name="Meiryo UI", size=11, color="000000", bold=False)
        header_font = Font(name="Meiryo UI", size=12, color="000000", bold=True)
        # 设置标题背景色
        header_bg_color = PatternFill(fill_type="solid", start_color="92D050", end_color="92D050")
        # 格式化第一行(标题行)
        for column_i in range(1, max_column + 1):
            work_sheet.cell(row=1, column=column_i).font = header_font
            work_sheet.cell(row=1, column=column_i).border = border
            work_sheet.cell(row=1, column=column_i).fill = header_bg_color
            max_column_len[column_i] = (
                max_column_len[column_i]
                if len(str(work_sheet.cell(row=1, column=column_i).value)) < max_column_len[column_i]
                else len(str(work_sheet.cell(row=1, column=column_i).value))
            )
        max_column_len = [x * 3 for x in max_column_len]
        # 格式化正文
        for row_i in range(2, max_row + 1):
            for column_i in range(1, max_column + 1):
                work_sheet.cell(row=row_i, column=column_i).font = content_font
                work_sheet.cell(row=row_i, column=column_i).border = border
                max_column_len[column_i] = (
                    max_column_len[column_i]
                    if len(str(work_sheet.cell(row=row_i, column=column_i).value)) * 1.28 + 5 < max_column_len[column_i]
                    else len(str(work_sheet.cell(row=row_i, column=column_i).value)) * 1.28 + 5
                )
        # 设置列宽
        for column_i in range(1, max_column + 1):
            work_sheet.column_dimensions[openpyxl.utils.get_column_letter(column_i)].width = max_column_len[column_i]
        # 筛选和固定标题行
        work_sheet.auto_filter.ref = work_sheet.dimensions
        work_sheet.freeze_panes = "A2"
    # Save
    work_book.save(excel_path)
    print("Formatting Finished")


def get_file_list(path, glob_type="file", recursive=True):
    assert os.path.exists(path), "file path is invalid : " + path
    files = glob.glob(path + "/**", recursive=recursive)
    if glob_type == "file":
        return [p for p in files if os.path.isfile(p)]
    elif glob_type == "folder":
        return [p for p in files if os.path.isdir(p)]


def parse_split_line(line: str) -> list:
    _ret = []
    start = 0
    for index, char in enumerate(line):
        if char == "-":
            if start == 0:
                start = index
        elif char == " " or char == "\n":
            if not start == 0:
                _ret.append([start, index])
                start = 0
    return _ret


def layout_parse(res_index: int, res: list, dbio: str, file_id: str, view_name: str, split_parm: list, line: str):
    _tyl = line[split_parm[TYL][0] : split_parm[TYL][1]]
    _name1 = line[split_parm[DB][0] : split_parm[DB][1]].strip()
    _name2 = line[split_parm[NAME][0] : split_parm[NAME][1]].strip()
    _format = line[split_parm[F][0] : split_parm[F][1]].strip()
    _leng = line[split_parm[LENG][0] : split_parm[LENG][1]].strip()
    _zip = line[split_parm[S][0] : split_parm[S][1]].strip()
    _disk = line[split_parm[D][0] : split_parm[D][1]].strip()
    # _remarks = line[split_parm[REMARKS][0]:split_parm[REMARKS][1]].strip()
    _group = _tyl[0].strip()
    _level = _tyl[-1].strip()
    res.append([res_index, dbio, file_id, view_name, _level, _group, _name1, _name2, _format, _leng, _zip, _disk])
    pass


def main(ddm_path: str, res_path: str):
    output_excel = os.path.join(res_path, "【CoE】XXXXX_NATURAL-DDMレイアウト定義.xlsx")
    res = []
    res_index = 1
    ddm_files = get_file_list(ddm_path)
    for ddm_file in tqdm.tqdm(ddm_files):
        dbio = file_id = view_name = ""
        split_parm = []
        with open(ddm_file, "r", encoding="CP932") as fr:
            lines = fr.readlines()
            # Skip comment line
            lines = [line for line in lines if not PATTERN_COMMENT.match(line)]
            # For each every line
            for line in lines:
                # Get DBIO, FILEID, VIEW_NAME
                if not line.startswith(" "):
                    m1 = PATTERN_F.match(line)
                    m2 = PATTERN_DB_FILE.match(line)
                    if m1:
                        view_name = m1.group(1)
                    elif m2:
                        dbio = m2.group(1)
                        file_id = m2.group(2)
                # Layout Parse
                else:
                    #  "---  --  -------------------------------- - ----  - -  ------------------------"
                    if PATTERN_LEN_LINE.match(line):
                        split_parm = parse_split_line(line)
                    #  "  1  KA  PART-NO                          A   12  F D                          "
                    else:
                        layout_parse(res_index, res, dbio, file_id, view_name, split_parm, line)
                        res_index = res_index + 1
    write_excel(output_excel, res, EXCEL_TITLE)


if __name__ == "__main__":
    # Release
    main(sys.argv[1], sys.argv[2])

    # Debug
    # i_path = r"C:\Users\yi.a.qian\Desktop\Work\231207_Natural_Layout_2_Excel\NATURAL_DDM"
    # o_path = r"C:\Users\yi.a.qian\Desktop\Work\231207_Natural_Layout_2_Excel\RES"
    # main(i_path, o_path)

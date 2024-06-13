import re
import sys
import os
import glob


import tqdm
import pandas
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side


EXCEL_TITLE = ["No", "View名", "レベル", "グループ", "物理名", "論理名", "フォーマット", "長さ", "繰り返し回数", "ディスクリプタ"]
CNT, T, L, FIELD_NAME, F, LENGTH, D, DB, OFF_DEC, OFF_HEX, OCC, GS, AL = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12

PATTERN_COMMENT = re.compile(r"^\s*\*")
PATTERN_VIEW_NAME = re.compile(r"\s*DATA\s+LIST\s+LAYOUT\s+(\S+?)[ ]")
PATTERN_LEN_LINE = re.compile(r"^\s*-")
PATTERN_LAYOUT = re.compile(r"^\s+\d")


def write_excel(output_path: str, result_list: list, column: list, format=True, index_flg=False) -> None:
    pandas.DataFrame(result_list, columns=column).to_excel(output_path, index=index_flg, engine="xlsxwriter")
    if format:
        formatter_excel(output_path)


def formatter_excel(excel_path) -> None:
    print(f"Formatting Excel: {os.path.basename(excel_path)}...")
    work_book = openpyxl.load_workbook(excel_path)
    for work_sheet_name in work_book.sheetnames:
        print(f" Sheet -> {work_sheet_name}...")
        work_sheet = work_book[work_sheet_name]
        max_row = work_sheet.max_row
        max_column = work_sheet.max_column
        max_column_len = [0] * (max_column + 1)
        # 设置边框
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )
        # 设置字体
        content_font = Font(name="Meiryo UI", size=11, color="000000", bold=False)
        header_font = Font(name="Meiryo UI", size=12, color="000000", bold=True)
        # 设置标题背景色
        header_bg_color = PatternFill(fill_type="solid", start_color="92D050", end_color="92D050")
        # 格式化第一行(标题行)
        for column_i in range(1, max_column + 1):
            work_sheet.cell(row=1, column=column_i).font = header_font
            work_sheet.cell(row=1, column=column_i).border = border
            work_sheet.cell(row=1, column=column_i).fill = header_bg_color
            max_column_len[column_i] = (
                max_column_len[column_i]
                if len(str(work_sheet.cell(row=1, column=column_i).value)) < max_column_len[column_i]
                else len(str(work_sheet.cell(row=1, column=column_i).value))
            )
        max_column_len = [x * 3 for x in max_column_len]
        # 格式化正文
        for row_i in range(2, max_row + 1):
            for column_i in range(1, max_column + 1):
                work_sheet.cell(row=row_i, column=column_i).font = content_font
                work_sheet.cell(row=row_i, column=column_i).border = border
                max_column_len[column_i] = (
                    max_column_len[column_i]
                    if len(str(work_sheet.cell(row=row_i, column=column_i).value)) * 1.28 + 5 < max_column_len[column_i]
                    else len(str(work_sheet.cell(row=row_i, column=column_i).value)) * 1.28 + 5
                )
        # 设置列宽
        for column_i in range(1, max_column + 1):
            work_sheet.column_dimensions[openpyxl.utils.get_column_letter(column_i)].width = max_column_len[column_i]
        # 筛选和固定标题行
        work_sheet.auto_filter.ref = work_sheet.dimensions
        work_sheet.freeze_panes = "A2"
    # Save
    work_book.save(excel_path)
    print("Formatting Finished")


def get_file_list(path, glob_type="file", recursive=True):
    assert os.path.exists(path), "file path is invalid : " + path
    files = glob.glob(path + "/**", recursive=recursive)
    if glob_type == "file":
        return [p for p in files if os.path.isfile(p)]
    elif glob_type == "folder":
        return [p for p in files if os.path.isdir(p)]


def parse_split_line(line: str) -> list:
    _ret = []
    start = 0
    for index, char in enumerate(line):
        if char == "-":
            if start == 0:
                start = index
        elif char == " " or char == "\n":
            if not start == 0:
                _ret.append([start, index])
                start = 0
    return _ret


def layout_parse(res_index: int, view_name: str, res: list, split_parm: list, line: str) -> None:
    _level = line[split_parm[L][0] : split_parm[L][1]].strip()
    _group = line[split_parm[T][0] : split_parm[T][1]].strip()
    _name1 = line[split_parm[DB][0] : split_parm[DB][1]].strip()
    _name2 = line[split_parm[FIELD_NAME][0] : split_parm[FIELD_NAME][1]].strip()
    _format = line[split_parm[F][0] : split_parm[F][1]].strip()
    _length = line[split_parm[LENGTH][0] : split_parm[LENGTH][1]].strip()
    _occ = line[split_parm[OCC][0] : split_parm[OCC][1]].strip()
    _disk = line[split_parm[D][0] : split_parm[D][1]].strip()
    res.append([res_index, view_name, _level, _group, _name1, _name2, _format, _length, _occ, _disk])


def main(predict_path: str, res_path: str):
    output_excel = os.path.join(res_path, "【CoE】XXXXX_COBOL-PREDICTレイアウト定義.xlsx")
    res = []
    res_index = 1
    predict_files = get_file_list(predict_path)
    for predict_file in tqdm.tqdm(predict_files):
        view_name = ""
        with open(predict_file, "r", encoding="CP932") as fr:
            lines = fr.readlines()
            # lines = [line[1:] for line in lines]
            # Skip comment line
            lines = [line for line in lines if not PATTERN_COMMENT.match(line)]
            # For each every line
            for line in lines:
                # Get view name
                match_view_name = PATTERN_VIEW_NAME.match(line)
                if match_view_name:
                    view_name = match_view_name.group(1)
                # Parse Layout
                else:
                    if PATTERN_LEN_LINE.match(line):
                        split_parm = parse_split_line(line)
                    elif PATTERN_LAYOUT.match(line):
                        layout_parse(res_index, view_name, res, split_parm, line)
                        res_index = res_index + 1
    write_excel(output_excel, res, EXCEL_TITLE)


if __name__ == "__main__":
    # Release
    main(sys.argv[1], sys.argv[2])

    # Debug
    # i_path = r"C:\Users\yi.a.qian\Desktop\Work\231208_Natural_PREDICT_Layout_2_Excel\PREDICT"
    # o_path = r"C:\Users\yi.a.qian\Desktop\Work\231208_Natural_PREDICT_Layout_2_Excel\RES"
    # main(i_path, o_path)

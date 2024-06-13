#!/usr/bin/env python
# -*- coding: cp932 -*-

import sys
import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from argparse import ArgumentParser

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

from Common_analysis import *

hyperlink_font = Font(underline='single', color='0563C1')
hyperlink_some_fill = PatternFill(patternType="solid",fgColor="FFFF00")

def make_index_dic(files,get_module=True,rem_extension=True):
    
    index_dic = {}
    
    for file in files:
        file_name = os.path.split(file)[-1]
        if rem_extension:
            file_name = take_extensions(file_name)
        if get_module:
            file_name = file_name.split("%")[-1]
        
        if file_name not in index_dic:
            index_dic[file_name] = []
        index_dic[file_name].append(file)
        
    return index_dic

def main(excel_path,out_folder,cobol_files,jcl_files,proc_files,schema_files):

    cobol_index_dic = make_index_dic(cobol_files,True,True)
    jcl_index_dic = make_index_dic(jcl_files+proc_files,False,True)
    proc_index_dic = make_index_dic(proc_files,False,True)
    schema_index_dic = make_index_dic(schema_files,False,False)

    excel_wb = openpyxl.load_workbook(excel_path)
    
    JCL_ROW = "B"
    PGM_ROW = "G"
    PROC_ROW = "H"
    SYSIN_ROW = "I"
    pgm_cands_index = 26
    sysin_cands_index = 31
    for work_sheet in excel_wb.sheetnames:
        if "DSN使用箇所" not in work_sheet:
            continue
        
        dsn_ws = excel_wb[work_sheet]
        
        ld = dsn_ws.max_row
        for i in range(2,dsn_ws.max_row+1):
            print("\r","analysis finished",i,"/",ld,end="")
            id_jcl = JCL_ROW + str(i)
            jcl = dsn_ws[id_jcl].value
            jcl = take_extensions(jcl)
            if jcl in jcl_index_dic:
                # dsn_ws[id_jcl].hyperlink = jcl_index_dic[jcl][0]
                dsn_ws[id_jcl].font = hyperlink_font
                
            id_pgm = PGM_ROW + str(i)
            pgm = dsn_ws[id_pgm].value
            if pgm in cobol_index_dic:
                # dsn_ws[id_pgm].hyperlink = cobol_index_dic[pgm][0]
                full_pgm = cobol_index_dic[pgm][0]
                full_pgm = take_extensions(os.path.split(full_pgm)[-1])
                dsn_ws[id_pgm].value = full_pgm
                dsn_ws[id_pgm].font = hyperlink_font
                
                if len(cobol_index_dic[pgm]) > 1:
                    assert len(cobol_index_dic[pgm]) <= 5
                    
                    for s,path in enumerate(cobol_index_dic[pgm]):
                        full_pgm = path
                        full_pgm = take_extensions(os.path.split(full_pgm)[-1])
                        dsn_ws.cell(i,pgm_cands_index+s).value = full_pgm
                        # dsn_ws.cell(i,pgm_cands_index+s).hyperlink = path
                        dsn_ws.cell(i,pgm_cands_index+s).font = hyperlink_font     
                        dsn_ws.cell(i,pgm_cands_index+s).fill = hyperlink_some_fill
                                
            
            id_proc = PROC_ROW + str(i)
            proc = dsn_ws[id_proc].value
            if proc in proc_index_dic:
            #     dsn_ws[id_proc].hyperlink = proc_index_dic[proc][0]
                dsn_ws[id_proc].font = hyperlink_font
                
            id_pgm = SYSIN_ROW + str(i)
            pgm = dsn_ws[id_pgm].value
            if pgm in cobol_index_dic:
                full_pgm = cobol_index_dic[pgm][0]
                full_pgm = take_extensions(os.path.split(full_pgm)[-1])
                dsn_ws[id_pgm].value = full_pgm
                dsn_ws[id_pgm].font = hyperlink_font
                
                if len(cobol_index_dic[pgm]) > 1:
                    assert len(cobol_index_dic[pgm]) <= 5
                    
                    for s,path in enumerate(cobol_index_dic[pgm]):
                        full_pgm = path
                        full_pgm = take_extensions(os.path.split(full_pgm)[-1])
                        dsn_ws.cell(i,sysin_cands_index+s).value = full_pgm
                        dsn_ws.cell(i,sysin_cands_index+s).font = hyperlink_font     
                        dsn_ws.cell(i,sysin_cands_index+s).fill = hyperlink_some_fill
            
    
    LAYOUT_ROW = "C"
    layout_detail_sheetname = "マルチレイアウト解析結果_明細"
    layout_detail_ws = excel_wb[layout_detail_sheetname]
    for i in range(2,layout_detail_ws.max_row+1):
        id_layout = LAYOUT_ROW + str(i)
        layout = layout_detail_ws[id_layout].value
        if layout in schema_index_dic:
            # layout_detail_ws[id_layout].hyperlink = schema_index_dic[layout][0]
            layout_detail_ws[id_layout].font = hyperlink_font
            
    out_file = os.path.join(out_folder,"レイアウト解析結果サマリ_リンク追加後.xlsx")
    excel_wb._save(out_file)
    excel_wb.close()
    
    
def parser():
    argparser = ArgumentParser()
    argparser.add_argument("-e","--excelpath",type=str,required=True,help="input the path to excel file")
    argparser.add_argument("-o","--outfolder",type=str,required=True,help="input the path to output folder")
    argparser.add_argument("-c","--cobol",type=str,default="",help="input the path to COBOL files")
    argparser.add_argument("-j","--jcl",type=str,default="",help="input the path to JCL files")
    argparser.add_argument("-p","--proc",type=str,default="",help="input the path to PROC files")
    argparser.add_argument("-s","--schema",type=str,default="",help="input the path to SCHEMA files")
    
    return argparser.parse_args()

if __name__ == "__main__":
    
    args = parser()
    if args.cobol == "":
        cobol_files = []
    else:
        cobol_files = glob_files(args.cobol)
        
    if args.jcl == "":
        jcl_files = []
    else:
        jcl_files = glob_files(args.jcl)
    
    if args.proc == "":
        proc_files = []
    else:
        proc_files = glob_files(args.proc)
    
    if args.schema == "":
        schema_files = []
    else:
        schema_files = glob_files(args.schema)
        
    print(len(cobol_files),len(jcl_files),len(proc_files),len(schema_files))
    main(args.excelpath,args.outfolder,cobol_files,jcl_files,proc_files,schema_files)
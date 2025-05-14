#!/usr/bin/env python
# -*- coding: cp932 -*-

import sys
import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

from Relation_Analysis.relation_analysis_functions import *


hyperlink_font = Font(underline='single', color='0563C1')

def main(excel_path):
# def main(excel_path,jcl_db_path,member_list_path):
    
    excel_wb = openpyxl.load_workbook(excel_path)
    
    
#     df_member = pd.read_excel(member_list_path,header=1,sheet_name="メンバ一覧マージ版")
#     update_member_list(df_member)
    
#     conn_jcl = connect_accdb(jcl_db_path)
#     sql = "SELECT * FROM ①JCL_基本情報"
#     df = pd.read_sql(sql,conn_jcl)
#     df.replace("", np.nan,inplace=True)
#     df.replace("\\","$",inplace=True)
    
#     update_internal_proc_list(df[df.JCL分類 == "内部PROC"])
    

# ### get the list of which source is jcl cmd info

#     sql = "SELECT * FROM ①JCL_CMD情報"
#     df = pd.read_sql(sql,conn_jcl)
#     df.replace("", np.nan,inplace=True)
#     df.replace("\\","$",inplace=True)
#     df.sort_values(["元資産行情報"],inplace=True)
#     update_jobproc_list(df[df.DD_NAME == "JOBPROC"])
    
    # GR_ROW = "A"
    PROC_ROW = "H"
    DSN_ROW = "K"
    JCL_ROW = "B"
    dsn_index_dic = {}
    
    for work_sheet in excel_wb.sheetnames:
        if "DSN使用箇所" not in work_sheet:
            continue
        
        dsn_ws = excel_wb[work_sheet]
        for i in range(2,dsn_ws.max_row+1):
            # id_gr = GR_ROW + str(i)
            # gr = dsn_ws[id_gr].value
            
            id_jcl = JCL_ROW + str(i)
            jcl = dsn_ws[id_jcl].value
            
            id_dsn = DSN_ROW+str(i)
            dsn = dsn_ws[id_dsn].value
            if dsn != None and dsn not in dsn_index_dic:
                # print(dsn,gr)
                dsn_index_dic[dsn] = "#"+work_sheet+"!"+id_dsn
                
                
            # id_proc = PROC_ROW + str(i)
            # proc = dsn_ws[id_proc].value
            
            # if proc == None:
            #     continue
            
        
            # proc_full = update_proc_library(jcl,proc)
            # dsn_ws[id_proc].value = proc_full
            
    
    DSN_ROW = "B"
    GROUPID_ROW = "A"
    
    dsn_group_index_dic = {}
    layout_detail_sheetname = "マルチレイアウト解析結果_明細"
    layout_detail_ws = excel_wb[layout_detail_sheetname]
    for i in range(2,layout_detail_ws.max_row+1):
        # id_gr = GR_ROW + str(i)
        # gr = layout_detail_ws[id_gr].value
        
        id_dsn = DSN_ROW + str(i)
        dsn = layout_detail_ws[id_dsn].value
        # print(dsn)
        if dsn in dsn_index_dic:
            layout_detail_ws[id_dsn].hyperlink = dsn_index_dic[dsn]
            # print(dsn_index_dic[(dsn,gr)],dsn,gr)
            layout_detail_ws[id_dsn].font = hyperlink_font
        
        id_group = GROUPID_ROW + str(i)
        groupid = layout_detail_ws[id_group].value
        if groupid != None and groupid not in dsn_group_index_dic:
            dsn_group_index_dic[groupid] = "#" + layout_detail_sheetname + "!" + id_group
            
            
    GROUPID_ROW = "A"
    layout_summary_sheetname = "マルチレイアウト解析結果_サマリ"
    layout_summary_ws = excel_wb[layout_summary_sheetname]
    for i in range(2,layout_summary_ws.max_row+1):
        # id_gr = GR_ROW + str(i)
        # gr = layout_summary_ws[id_gr].value
        
        id_group = GROUPID_ROW + str(i)
        groupid = layout_summary_ws[id_group].value
        if groupid in dsn_group_index_dic:
            # link = excel_path +dsn_group_index_dic[groupid]
            link =dsn_group_index_dic[groupid]
            
            layout_summary_ws[id_group].hyperlink = link
            layout_summary_ws[id_group].font = hyperlink_font
        
            
    excel_wb._save(excel_path)
    
if __name__ == "__main__":
    main(sys.argv[1])
    
    # main(sys.argv[1],sys.argv[2],sys.argv[3])